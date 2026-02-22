"""Expense Excel Generation Module."""
import os
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Excel output directory
OUTPUT_DIR = os.path.expanduser("~/projects/bim/expenses/output")
os.makedirs(OUTPUT_DIR, exist_ok=True)


def generate_excel(month_key: str, receipts: List[Dict[str, Any]]) -> str:
    """
    Generate Excel file for a month's expenses.
    
    Args:
        month_key: YYYY-MM format
        receipts: List of receipt dictionaries
        
    Returns:
        Path to generated Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    # Headers
    headers = ["Date", "Description", "Category", "Original Amount", "Currency", "FX Rate", "Amount EUR"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill('solid', fgColor='D5E8F0')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Column widths
    ws.column_dimensions['A'].width = 12   # Date
    ws.column_dimensions['B'].width = 40   # Description
    ws.column_dimensions['C'].width = 16   # Category
    ws.column_dimensions['D'].width = 16   # Original Amount
    ws.column_dimensions['E'].width = 10   # Currency
    ws.column_dimensions['F'].width = 12   # FX Rate
    ws.column_dimensions['G'].width = 14   # Amount EUR
    
    # Sort receipts by date
    receipts_sorted = sorted(
        [r for r in receipts if r.get('status') == 'active'],
        key=lambda r: r.get('date', '')
    )
    
    # Data rows
    for i, receipt in enumerate(receipts_sorted):
        row = 2 + i
        
        # Build description from vendor + description
        vendor = receipt.get('vendor', '')
        desc = receipt.get('description', '')
        description = f"{vendor} - {desc}" if vendor and desc else (vendor or desc or 'N/A')
        
        ws.cell(row=row, column=1, value=receipt.get('date', ''))           # A: Date
        ws.cell(row=row, column=2, value=description)                         # B: Description
        ws.cell(row=row, column=3, value=receipt.get('category', ''))          # C: Category
        ws.cell(row=row, column=4, value=receipt.get('original_amount', 0))   # D: Original Amount
        ws.cell(row=row, column=5, value=receipt.get('original_currency', '')) # E: Currency
        ws.cell(row=row, column=6, value=receipt.get('fx_rate', 1))           # F: FX Rate
        ws.cell(row=row, column=7, value=receipt.get('amount_eur', 0))        # G: Amount EUR
        
        # Number formatting
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        ws.cell(row=row, column=7).number_format = '#,##0.00'
    
    # Totals row
    total_row = 2 + len(receipts_sorted)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=f'=SUM(G2:G{total_row - 1})')
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    
    # Save
    year, month = month_key.split('-')
    filename = f"expenses_{year}_{month}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    
    return filepath


def get_month_name(month_num: str) -> str:
    """Convert month number to name."""
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 
              'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    try:
        return months[int(month_num) - 1]
    except (ValueError, IndexError):
        return month_num


if __name__ == "__main__":
    # Test with sample data
    test_receipts = [
        {
            'date': '2026-02-01',
            'vendor': 'Uber',
            'description': 'Airport ride',
            'category': 'travel',
            'original_amount': 25.00,
            'original_currency': 'USD',
            'fx_rate': 0.92,
            'amount_eur': 23.00,
            'status': 'active'
        },
        {
            'date': '2026-02-05',
            'vendor': 'Restaurant El Huaso',
            'description': 'Team lunch',
            'category': 'food',
            'original_amount': 45.00,
            'original_currency': 'EUR',
            'fx_rate': 1.0,
            'amount_eur': 45.00,
            'status': 'active'
        }
    ]
    
    filepath = generate_excel('2026-02', test_receipts)
    print(f"Generated: {filepath}")
